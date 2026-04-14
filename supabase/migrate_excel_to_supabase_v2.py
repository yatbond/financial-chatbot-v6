#!/usr/bin/env python3
"""
Financial Chatbot v6 - Excel to Supabase Migration (v2 Schema)

Reads Excel files, syncs lookup tables from CSVs, and uploads to Supabase.

Schema v2 changes:
- report_year / report_month (from Report Date)
- data_month (NULL for Financial Status snapshots, 4-12 for monthly sheets)
- Lookup tables: financial_types, line_items, acronyms (synced from CSVs)
- UPSERT with value_change_log audit trail

Usage:
    python migrate_excel_to_supabase.py --input-dir "/mnt/g/My Drive/Ai Chatbot Knowledge Base/Processing"
    python migrate_excel_to_supabase.py --input-dir "/mnt/g/My Drive/Ai Chatbot Knowledge Base/Processing" --dry-run
"""

import os
import sys
import re
import shutil
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Tuple, Optional, Any

import pandas as pd
from openpyxl import load_workbook
from supabase import create_client, Client

# Support for old .xls files
try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False
    print("[WARN] xlrd not installed - .xls files will be skipped")

# ============================================================================
# Configuration
# ============================================================================

SUPABASE_URL = os.environ.get('SUPABASE_URL')
SUPABASE_KEY = os.environ.get('SUPABASE_KEY')  # Service role key preferred

if not SUPABASE_URL or not SUPABASE_KEY:
    print("ERROR: Set SUPABASE_URL and SUPABASE_KEY environment variables")
    print("  export SUPABASE_URL='https://xxx.supabase.co'")
    print("  export SUPABASE_KEY='your-service-role-key'")
    sys.exit(1)

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

PROCESSING_DIR = Path(r'G:\My Drive\Ai Chatbot Knowledge Base\Processing')
FINANCIAL_TYPE_CSV = PROCESSING_DIR / 'financial_type_map.csv'
HEADINGS_CSV = PROCESSING_DIR / 'construction_headings_enriched.csv'

# Month name mapping
MONTH_MAP = {
    'jan': 1, 'january': 1, 'feb': 2, 'february': 2, 'mar': 3, 'march': 3,
    'apr': 4, 'april': 4, 'may': 5, 'jun': 6, 'june': 6, 'jul': 7, 'july': 7,
    'aug': 8, 'august': 8, 'sep': 9, 'september': 9, 'oct': 10, 'october': 10,
    'nov': 11, 'november': 11, 'dec': 12, 'december': 12,
}

# ============================================================================
# Step 1: Sync Lookup Tables from CSVs
# ============================================================================

def sync_financial_types():
    """Sync financial_types table from financial_type_map.csv."""
    if not FINANCIAL_TYPE_CSV.exists():
        print(f"  [WARN] {FINANCIAL_TYPE_CSV} not found, skipping financial_types sync")
        return

    df = pd.read_csv(FINANCIAL_TYPE_CSV)
    synced = 0

    # Sheet name mapping: which clean_name maps to which Excel sheet
    sheet_name_map = {
        'Budget Tender': 'Financial Status',
        '1st Working Budget': 'Financial Status',
        'Latest Budget': 'Financial Status',
        'Business Plan': 'Financial Status',
        'WIP': 'Financial Status',
        'Projection': 'Projection',
        'Committed Cost': 'Committed Cost',
        'Accrual': 'Accrual',
        'Cash Flow': 'Cash Flow',
        'General': 'Financial Status',
    }

    for _, row in df.iterrows():
        clean_name = str(row.get('Clean_Financial_Type', '')).strip()
        if not clean_name or clean_name == '*not used':
            continue

        raw_type = str(row.get('Raw_Financial_Type', '')).strip()
        acronyms_str = str(row.get('Acronyms', '')).strip()

        # Collect all raw names for this clean name
        raw_names = [raw_type] if raw_type and raw_type != clean_name else []
        # Also collect acronyms
        acronyms = [a.strip() for a in acronyms_str.split('|') if a.strip()] if acronyms_str else []
        sheet_name = sheet_name_map.get(clean_name, 'Financial Status')

        # Upsert financial_types
        data = {
            'clean_name': clean_name,
            'raw_names': raw_names,
            'acronyms': acronyms,
            'sheet_name': sheet_name,
            'is_active': True,
        }

        try:
            # Check if exists
            existing = supabase.table('financial_types').select('id').eq('clean_name', clean_name).execute()
            if existing.data:
                supabase.table('financial_types').update(data).eq('clean_name', clean_name).execute()
            else:
                supabase.table('financial_types').insert(data).execute()
            synced += 1
        except Exception as e:
            print(f"  [ERROR] Failed to sync financial_type '{clean_name}': {e}")

    print(f"  [OK] Synced {synced} financial_types")

    # Sync acronyms table
    _sync_acronyms_from_ft(df)


def _sync_acronyms_from_ft(df):
    """Sync acronyms for financial types — batch mode."""
    # Fetch all existing financial_type acronyms
    existing = supabase.table('acronyms').select('acronym').eq('target_table', 'financial_type').execute()
    existing_set = {(r['acronym']) for r in existing.data} if existing.data else set()

    to_insert = []
    for _, row in df.iterrows():
        clean_name = str(row.get('Clean_Financial_Type', '')).strip()
        if not clean_name or clean_name == '*not used':
            continue
        acronyms_str = str(row.get('Acronyms', '')).strip()
        if not acronyms_str:
            continue
        for acronym in acronyms_str.split('|'):
            acronym = acronym.strip()
            if acronym and acronym not in existing_set:
                to_insert.append({'target_table': 'financial_type', 'target_id': clean_name, 'acronym': acronym})
                existing_set.add(acronym)

    if to_insert:
        for i in range(0, len(to_insert), 50):
            supabase.table('acronyms').insert(to_insert[i:i+50]).execute()
    print(f"  [OK] Synced {len(to_insert)} financial_type acronyms")


def sync_line_items():
    """Sync line_items table from construction_headings_enriched.csv."""
    if not HEADINGS_CSV.exists():
        print(f"  [WARN] {HEADINGS_CSV} not found, skipping line_items sync")
        return

    df = pd.read_csv(HEADINGS_CSV)
    synced = 0
    acronym_count = 0
    acronym_rows = []  # Collect acronym inserts for batch

    for _, row in df.iterrows():
        item_code = str(row.get('Item_Code', '')).strip()
        data_type = str(row.get('Data_Type', '')).strip() if pd.notna(row.get('Data_Type')) else ''
        friendly_name = str(row.get('Friendly_Name', '')).strip() if pd.notna(row.get('Friendly_Name')) else data_type
        category = str(row.get('Category', '')).strip() if pd.notna(row.get('Category')) else ''
        tier = int(row['Tier']) if pd.notna(row.get('Tier')) else 0
        acronyms_str = str(row.get('Acronyms', '')).strip() if pd.notna(row.get('Acronyms')) else ''

        # Determine parent code
        parent_code = None
        if item_code and '.' in item_code:
            parts = item_code.rsplit('.', 1)
            parent_code = parts[0]

        # For Project Info items (no item_code), use Data_Type as key
        pk = item_code if item_code else data_type
        if not pk:
            continue

        # Upsert line_items
        data = {
            'item_code': pk,
            'data_type': data_type,
            'friendly_name': friendly_name,
            'category': category,
            'tier': tier,
            'acronyms': [a.strip() for a in acronyms_str.split('|') if a.strip()] if acronyms_str else [],
            'parent_code': parent_code,
        }

        try:
            existing = supabase.table('line_items').select('item_code').eq('item_code', pk).execute()
            if existing.data:
                supabase.table('line_items').update(data).eq('item_code', pk).execute()
            else:
                supabase.table('line_items').insert(data).execute()
            synced += 1
        except Exception as e:
            print(f"  [ERROR] Failed to sync line_item '{pk}': {e}")

        if acronyms_str:
            for acronym in acronyms_str.split('|'):
                acronym = acronym.strip()
                if acronym:
                    acronym_rows.append({'target_table': 'line_item', 'target_id': pk, 'acronym': acronym})
    if acronym_rows:
        existing = supabase.table('acronyms').select('target_id, acronym').eq('target_table', 'line_item').execute()
        existing_set = {(r['target_id'], r['acronym']) for r in existing.data} if existing.data else set()
        to_insert = [r for r in acronym_rows if (r['target_id'], r['acronym']) not in existing_set]
        if to_insert:
            for i in range(0, len(to_insert), 50):
                supabase.table('acronyms').insert(to_insert[i:i+50]).execute()
        acronym_count = len(to_insert)
    else:
        acronym_count = 0

    print(f"  [OK] Synced {synced} line_items, {acronym_count} acronyms")


def sync_lookup_tables():
    """Sync all lookup tables from CSVs."""
    print("\n[STEP 1] Syncing lookup tables from CSVs...")
    sync_financial_types()
    sync_line_items()


# ============================================================================
# Step 2: Parse Excel Files
# ============================================================================

# Load heading lookup for parsing
_heading_lookup: Dict[str, Tuple[str, str, str]] = {}  # Data_Type -> (Item_Code, Friendly_Name, Category)
_heading_by_code: Dict[str, Tuple[str, str]] = {}  # Item_Code -> (Friendly_Name, Category)
_financial_type_lookup: Dict[str, str] = {}  # Raw -> Clean


def load_local_lookups():
    """Load CSV lookups into memory for parsing."""
    global _heading_lookup, _heading_by_code, _financial_type_lookup

    if HEADINGS_CSV.exists():
        df = pd.read_csv(HEADINGS_CSV)
        for _, row in df.iterrows():
            data_type = str(row.get('Data_Type', '')).strip() if pd.notna(row.get('Data_Type')) else ''
            item_code = str(row.get('Item_Code', '')).strip() if pd.notna(row.get('Item_Code')) else ''
            friendly_name = str(row.get('Friendly_Name', '')).strip() if pd.notna(row.get('Friendly_Name')) else data_type
            category = str(row.get('Category', '')).strip() if pd.notna(row.get('Category')) else ''

            if data_type:
                _heading_lookup[data_type] = (item_code, friendly_name, category)
            if item_code:
                _heading_by_code[item_code] = (friendly_name, category)

    if FINANCIAL_TYPE_CSV.exists():
        df = pd.read_csv(FINANCIAL_TYPE_CSV)
        for _, row in df.iterrows():
            raw = str(row.get('Raw_Financial_Type', '')).strip()
            clean = str(row.get('Clean_Financial_Type', '')).strip()
            if clean == '*not used' or not clean:
                continue
            if raw:
                _financial_type_lookup[raw] = clean
            if clean:
                _financial_type_lookup[clean] = clean
            acronyms = str(row.get('Acronyms', '')).strip()
            for a in acronyms.split('|'):
                a = a.strip()
                if a:
                    _financial_type_lookup[a] = clean


def translate_data_type(raw: str) -> Tuple[str, str, str, str]:
    """Translate raw Data_Type to (Item_Code, Friendly_Name, Category, Match_Status)."""
    if not raw:
        return '', raw, '', 'UNMAPPED'
    raw = str(raw).strip()

    if raw in _heading_by_code:
        fn, cat = _heading_by_code[raw]
        return raw, fn, cat, 'EXACT'
    if raw in _heading_lookup:
        ic, fn, cat = _heading_lookup[raw]
        return ic, fn, cat, 'EXACT'

    # Fuzzy
    import difflib
    matches = difflib.get_close_matches(raw, _heading_lookup.keys(), n=1, cutoff=0.90)
    if matches:
        ic, fn, cat = _heading_lookup[matches[0]]
        return ic, fn, cat, 'FUZZY'

    return '', raw, '', 'UNMAPPED'


def translate_financial_type(raw: str) -> str:
    """Translate raw Financial_Type to clean name."""
    if not raw:
        return raw
    raw = str(raw).strip()
    if raw in _financial_type_lookup:
        return _financial_type_lookup[raw]

    import difflib
    matches = difflib.get_close_matches(raw, _financial_type_lookup.keys(), n=1, cutoff=0.90)
    if matches:
        return _financial_type_lookup[matches[0]]
    return raw


def get_report_date(ws) -> Tuple[int, int]:
    """Extract report date from sheet header."""
    for row_idx in range(1, 15):
        cell_a = ws.cell(row_idx, 1).value
        if cell_a and 'Report Date' in str(cell_a):
            cell_b = ws.cell(row_idx, 2).value
            if cell_b:
                date_str = str(cell_b).strip()
                if '  ' in date_str:
                    date_str = date_str.split('  ')[0].strip()
                try:
                    dt = pd.to_datetime(date_str)
                    return dt.year, dt.month
                except (ValueError, TypeError):
                    pass
    return 2025, 1


def find_header_row(ws) -> Optional[int]:
    """Find header row containing 'Item'."""
    for row_idx in range(1, min(ws.max_row + 1, 30)):
        for col in [1, 2]:
            val = ws.cell(row_idx, col).value
            if val and str(val).strip().lower() == 'item':
                return row_idx
    return None


def parse_financial_status(ws, header_row: int, report_year: int, report_month: int,
                           project_code: str, project_name: str, source_file: str) -> List[Dict]:
    """Parse Financial Status sheet (snapshot, no monthly columns)."""
    rows = []

    # Build column → financial_type mapping
    col_map = {}
    for col in range(2, ws.max_column):
        parts = []
        for row_offset in range(0, 4):
            val = ws.cell(header_row + row_offset, col + 1).value
            if val:
                parts.append(str(val).strip())
        combined = ' '.join(parts)
        if combined:
            clean = translate_financial_type(combined)
            if clean != '*not used':
                col_map[col] = clean

    trade_col = 2  # Column B = trade name

    for row_idx in range(header_row + 5, ws.max_row + 1):
        item_val = ws.cell(row_idx, 1).value
        if not item_val:
            continue

        item_str = str(item_val).strip()
        # Normalize item code (e.g. "1.0" → "1")
        if '.' in item_str:
            parts = item_str.split('.')
            if len(parts) == 2 and parts[1] == '0':
                item_str = parts[0]

        trade_str = str(ws.cell(row_idx, trade_col).value or '').strip()
        if not trade_str or trade_str == 'None':
            continue

        item_code, friendly_name, category, match_status = translate_data_type(trade_str)

        for col_idx, clean_ft in col_map.items():
            if col_idx + 1 > ws.max_column:
                continue
            value = ws.cell(row_idx, col_idx + 1).value
            if value is None or value == '':
                continue

            rows.append({
                'project_code': project_code,
                'project_name': project_name,
                'report_year': report_year,
                'report_month': report_month,
                'data_month': None,  # Snapshot — no monthly breakdown
                'financial_type': clean_ft,
                'raw_financial_type': clean_ft,
                'item_code': item_code if item_code else item_str,
                'friendly_name': friendly_name,
                'category': category,
                'value': float(value) if isinstance(value, (int, float)) else None,
                'raw_value': str(value),
                'match_status': match_status,
                'source_file': source_file,
            })

    return rows


def parse_monthly_sheet(ws, header_row: int, report_year: int, report_month: int,
                        financial_type: str, project_code: str, project_name: str,
                        source_file: str) -> List[Dict]:
    """Parse monthly sheet (Cash Flow, Projection, Accrual, Committed Cost).
    Each column = a data_month."""
    rows = []

    # Safely get max_column (handle corrupted cells)
    try:
        max_col = ws.max_column
    except (TypeError, AttributeError):
        # Fallback: scan for last non-empty column in header row
        max_col = 1
        for c in range(1, 50):
            if ws.cell(header_row, c).value is not None:
                max_col = c

    # Build month column mapping
    month_cols = {}
    for col in range(0, max_col):
        val = ws.cell(header_row, col + 1).value
        if val:
            if hasattr(val, 'month'):
                month_cols[val.month] = col
            else:
                header = str(val).strip()
                if header.lower() in MONTH_MAP:
                    month_cols[MONTH_MAP[header.lower()]] = col

    trade_col = 2  # Column B = trade name

    for row_idx in range(header_row + 1, ws.max_row + 1):
        item_val = ws.cell(row_idx, 1).value
        if not item_val:
            continue

        item_str = str(item_val).strip()
        if '.' in item_str:
            parts = item_str.split('.')
            if len(parts) == 2 and parts[1] == '0':
                item_str = parts[0]

        trade_str = str(ws.cell(row_idx, trade_col).value or '').strip()
        if not trade_str or trade_str == 'None':
            continue

        item_code, friendly_name, category, match_status = translate_data_type(trade_str)

        for month_num, col_idx in month_cols.items():
            if col_idx + 1 > max_col:
                continue
            value = ws.cell(row_idx, col_idx + 1).value
            if value is None or value == '':
                continue

            rows.append({
                'project_code': project_code,
                'project_name': project_name,
                'report_year': report_year,
                'report_month': report_month,
                'data_month': month_num,  # The month this data point belongs to
                'financial_type': financial_type,
                'raw_financial_type': financial_type,
                'item_code': item_code if item_code else item_str,
                'friendly_name': friendly_name,
                'category': category,
                'value': float(value) if isinstance(value, (int, float)) else None,
                'raw_value': str(value),
                'match_status': match_status,
                'source_file': source_file,
            })

    return rows


def parse_excel_file(excel_path: Path) -> List[Dict]:
    """Parse an Excel file into rows for the v2 schema."""
    # Handle .xls files (old format)
    if excel_path.suffix.lower() == '.xls':
        if not HAS_XLRD:
            print(f"  [SKIP] .xls file requires xlrd: {excel_path.name}")
            return []
        try:
            # Use pandas to read .xls, convert to intermediate format
            xls = pd.ExcelFile(excel_path, engine='xlrd')
            # Convert to temporary .xlsx-like structure using openpyxl
            # Actually, let's just parse with pandas directly
            return parse_excel_with_pandas(excel_path, xls)
        except Exception as e:
            print(f"  [ERROR] Failed to parse .xls {excel_path.name}: {e}")
            return []
    
    # Handle .xlsx files (new format)
    try:
        wb = load_workbook(excel_path, data_only=True)
    except Exception as e:
        print(f"  [ERROR] Failed to load {excel_path}: {e}")
        return []

    # Extract report date from first sheet
    report_year, report_month = 2025, 1
    for sn in wb.sheetnames:
        ws = wb[sn]
        report_year, report_month = get_report_date(ws)
        if report_year > 2025:
            break

    # Extract project code/name from filename
    filename = excel_path.stem
    code_match = re.match(r'^(\d+)\s*(.+)?', filename)
    project_code = code_match.group(1) if code_match else '0000'
    project_name = code_match.group(2).strip() if code_match and code_match.group(2) else filename

    all_rows = []

    # Financial type per sheet
    sheet_ft_map = {
        'cash flow': 'Cash Flow',
        'cashflow': 'Cash Flow',
        'projection': 'Projection',
        'projected cost': 'Projection',
        'committed cost': 'Committed Cost',
        'committed': 'Committed Cost',
        'accrual': 'Accrual',
        "actual rec'd & cost": 'Accrual',
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 5:
            continue

        header_row = find_header_row(ws)
        if header_row is None:
            continue

        sheet_lower = sheet_name.lower().strip()

        if 'financial status' in sheet_lower:
            rows = parse_financial_status(
                ws, header_row, report_year, report_month,
                project_code, project_name, excel_path.name
            )
        else:
            # Determine financial type
            ft = None
            for key, val in sheet_ft_map.items():
                if key in sheet_lower:
                    ft = val
                    break
            if not ft:
                continue  # Skip unknown sheets

            rows = parse_monthly_sheet(
                ws, header_row, report_year, report_month,
                ft, project_code, project_name, excel_path.name
            )

        all_rows.extend(rows)

    return all_rows


# ============================================================================
# Step 3: UPSERT to Supabase with Change Logging
# ============================================================================

def upsert_project(project_code: str, project_name: str) -> str:
    """Insert or get project, return project ID."""
    result = supabase.table('projects').select('id, name').eq('code', project_code).execute()

    if result.data and len(result.data) > 0:
        # Update name if changed
        existing = result.data[0]
        if existing['name'] != project_name:
            supabase.table('projects').update({'name': project_name}).eq('id', existing['id']).execute()
        return existing['id']

    result = supabase.table('projects').insert({
        'code': project_code,
        'name': project_name,
    }).execute()

    return result.data[0]['id']


def upsert_rows(rows: List[Dict], batch_size: int = 250):
    """UPSERT rows to Supabase with change logging — batch mode."""
    if not rows:
        return

    # Group by project
    projects: Dict[str, List[Dict]] = {}
    for row in rows:
        projects.setdefault(row['project_code'], []).append(row)

    total_upserted = 0
    total_changes = 0

    for project_code, project_rows in projects.items():
        project_name = project_rows[0]['project_name']
        project_id = upsert_project(project_code, project_name)
        print(f"  Project {project_code} ({project_name}): {len(project_rows)} rows")

        # Fetch all existing rows for this project in one query
        existing_data = supabase.table('financial_data').select('id, project_id, report_year, report_month, data_month, financial_type, item_code, value, source_file').eq('project_id', project_id).execute()
        existing_map = {}
        for r in (existing_data.data or []):
            key = (r['report_year'], r['report_month'], r.get('data_month'), r['financial_type'], r['item_code'])
            existing_map[key] = r

        to_upsert = []
        change_logs = []
        seen_keys = set()  # Deduplicate within this project (latest file wins)

        for row in project_rows:
            key = (row['report_year'], row['report_month'], row['data_month'], row['financial_type'], row['item_code'])
            
            # Skip duplicates within same project (latest file wins - already sorted by filename)
            if key in seen_keys:
                continue
            seen_keys.add(key)
            
            upsert_data = {
                'project_id': project_id,
                'report_year': row['report_year'],
                'report_month': row['report_month'],
                'data_month': row['data_month'],
                'financial_type': row['financial_type'],
                'item_code': row['item_code'],
                'value': row['value'],
                'raw_value': row['raw_value'],
                'match_status': row['match_status'],
                'source_file': row['source_file'],
                'friendly_name': row.get('friendly_name', ''),
                'category': row.get('category', ''),
            }

            if key in existing_map:
                # Row exists — check if value changed
                old = existing_map[key]
                old_value = old.get('value')
                new_value = row['value']

                values_differ = False
                if old_value is None and new_value is not None:
                    values_differ = True
                elif old_value is not None and new_value is None:
                    values_differ = True
                elif old_value is not None and new_value is not None:
                    try:
                        values_differ = abs(float(old_value) - float(new_value)) > 0.001
                    except (ValueError, TypeError):
                        values_differ = str(old_value) != str(new_value)

                if values_differ:
                    change_logs.append({
                        'project_id': project_id,
                        'report_year': row['report_year'],
                        'report_month': row['report_month'],
                        'data_month': row['data_month'],
                        'financial_type': row['financial_type'],
                        'item_code': row['item_code'],
                        'old_value': old_value,
                        'new_value': new_value,
                        'old_source': old.get('source_file', ''),
                        'new_source': row['source_file'],
                    })

            to_upsert.append(upsert_data)

        # Batch upsert all rows
        if to_upsert:
            for i in range(0, len(to_upsert), batch_size):
                batch = to_upsert[i:i+batch_size]
                (supabase.table('financial_data')
                    .upsert(batch, on_conflict='project_id,report_year,report_month,data_month,financial_type,item_code')
                    .execute())
            total_upserted += len(to_upsert)

        # Batch insert change logs
        if change_logs:
            for i in range(0, len(change_logs), batch_size):
                supabase.table('value_change_log').insert(change_logs[i:i+batch_size]).execute()
            total_changes += len(change_logs)

    print(f"\n  [OK] Upserted {total_upserted} rows, logged {total_changes} value changes")


# ============================================================================
# Main
# ============================================================================

def find_excel_files(input_dir: Path) -> List[Path]:
    """Find all Excel files in directory (top-level only, skip subfolders like 'processed')."""
    excel_files = []
    for ext in ['*.xlsx', '*.xls']:
        excel_files.extend(input_dir.glob(ext))
    return sorted(excel_files)


def main():
    import argparse
    parser = argparse.ArgumentParser(description='Migrate Excel files to Supabase (v2 schema)')
    parser.add_argument('--input-dir', required=True, help='Input directory with Excel files')
    parser.add_argument('--dry-run', action='store_true', help="Parse but don't upload")
    parser.add_argument('--skip-lookups', action='store_true', help="Skip lookup table sync")
    args = parser.parse_args()

    input_dir = Path(args.input_dir)
    if not input_dir.exists():
        print(f"ERROR: Directory not found: {input_dir}")
        sys.exit(1)

    print(f"[START] Excel to Supabase migration (v2 schema)")
    print(f"  Input: {input_dir}")
    print(f"  Supabase: {SUPABASE_URL}")
    print(f"  Dry run: {args.dry_run}")

    # Step 1: Sync lookup tables
    if not args.skip_lookups:
        sync_lookup_tables()

    # Load local lookups for parsing
    load_local_lookups()

    # Step 2: Find and parse Excel files
    excel_files = find_excel_files(input_dir)
    print(f"\n[STEP 2] Found {len(excel_files)} Excel files")

    if not excel_files:
        print("[WARN] No Excel files found.")
        return

    processed_dir = input_dir / 'processed'
    processed_dir.mkdir(exist_ok=True)

    all_rows = []
    file_results = []  # (path, row_count)
    for i, excel_path in enumerate(excel_files, 1):
        print(f"\n  [{i}/{len(excel_files)}] {excel_path.name}")
        rows = parse_excel_file(excel_path)
        print(f"    Parsed {len(rows)} rows")
        all_rows.extend(rows)
        file_results.append((excel_path, len(rows)))

    print(f"\n[SUMMARY] Total: {len(all_rows)} rows from {len(excel_files)} files")

    # Step 3: UPSERT to Supabase
    if args.dry_run:
        print("\n[DRY RUN] Not uploading to Supabase")
    else:
        print("\n[STEP 3] Upserting to Supabase with change logging...")
        upsert_rows(all_rows)

        # Step 4: Move successfully processed files to processed/ folder
        print("\n[STEP 4] Moving processed files...")
        moved = 0
        for excel_path, row_count in file_results:
            if row_count > 0:
                dest = processed_dir / excel_path.name
                if dest.exists():
                    dest.unlink()  # Remove existing
                shutil.move(str(excel_path), str(dest))
                print(f"  Moved: {excel_path.name} -> processed/")
                moved += 1
            else:
                print(f"  Skipped (0 rows): {excel_path.name}")
        print(f"  Moved {moved} files to processed/")

    print("\n[DONE]")


if __name__ == '__main__':
    main()
