#!/usr/bin/env python3
"""
Financial Chatbot v6 - Excel to Supabase Migration

Reads Excel files directly from Google Drive folder and uploads to Supabase.
Uses the same parsing logic as financial_preprocessor_v5.py.

Usage:
    python migrate_excel_to_supabase.py --input-dir "/mnt/g/My Drive/Ai Chatbot Knowledge Base/Processing"
    python migrate_excel_to_supabase.py --input-dir "/mnt/g/My Drive/Ai Chatbot Knowledge Base/2026/2"
"""

import os
import sys
import re
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Tuple, Optional, Any

# Add Processing folder to path to import preprocessor
sys.path.insert(0, '/mnt/g/My Drive/Ai Chatbot Knowledge Base/Processing')

import pandas as pd
from openpyxl import load_workbook
from supabase import create_client, Client

# ============================================================================
# Configuration
# ============================================================================

SUPABASE_URL = os.environ.get('SUPABASE_URL')
SUPABASE_KEY = os.environ.get('SUPABASE_KEY')

if not SUPABASE_URL or not SUPABASE_KEY:
    print("ERROR: Set SUPABASE_URL and SUPABASE_KEY environment variables")
    print("  export SUPABASE_URL='https://xxx.supabase.co'")
    print("  export SUPABASE_KEY='your-service-role-key'")
    sys.exit(1)

# Initialize Supabase client
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# ============================================================================
# Lookup Table Loading (reuse preprocessor logic)
# ============================================================================

heading_lookup: Dict[str, Tuple[str, str, str]] = {}  # Data_Type -> (Item_Code, Friendly_Name, Category)
heading_by_code: Dict[str, Tuple[str, str]] = {}  # Item_Code -> (Friendly_Name, Category)
financial_type_lookup: Dict[str, str] = {}  # Raw -> Clean

def load_lookup_tables():
    """Load heading and financial type mappings from CSV files."""
    global heading_lookup, heading_by_code, financial_type_lookup
    
    processing_dir = Path('/mnt/g/My Drive/Ai Chatbot Knowledge Base/Processing')
    
    # Load heading mapping
    heading_csv = processing_dir / 'construction_headings_enriched.csv'
    if heading_csv.exists():
        df = pd.read_csv(heading_csv)
        for _, row in df.iterrows():
            data_type = str(row['Data_Type']).strip() if pd.notna(row['Data_Type']) else ''
            item_code = str(row['Item_Code']).strip() if pd.notna(row['Item_Code']) else ''
            friendly_name = str(row['Friendly_Name']).strip() if pd.notna(row['Friendly_Name']) else data_type
            category = str(row['Category']).strip() if pd.notna(row['Category']) else ''
            
            if data_type:
                heading_lookup[data_type] = (item_code, friendly_name, category)
            if item_code:
                heading_by_code[item_code] = (friendly_name, category)
        print(f"[OK] Loaded {len(heading_lookup)} heading mappings")
    
    # Load financial type mapping
    ft_csv = processing_dir / 'financial_type_map.csv'
    if ft_csv.exists():
        df = pd.read_csv(ft_csv)
        for _, row in df.iterrows():
            raw_type = str(row['Raw_Financial_Type']).strip() if pd.notna(row['Raw_Financial_Type']) else ''
            clean_type = str(row['Clean_Financial_Type']).strip() if pd.notna(row['Clean_Financial_Type']) else raw_type
            acronyms = str(row['Acronyms']).strip() if pd.notna(row['Acronyms']) else ''
            
            if clean_type == '*not used':
                continue
            
            if raw_type:
                financial_type_lookup[raw_type] = clean_type
            if clean_type:
                financial_type_lookup[clean_type] = clean_type
            if acronyms:
                for acronym in acronyms.split('|'):
                    acronym = acronym.strip()
                    if acronym:
                        financial_type_lookup[acronym] = clean_type
        print(f"[OK] Loaded {len(financial_type_lookup)} financial type mappings")

def translate_data_type(raw_data_type: str) -> Tuple[str, str, str, str]:
    """Translate raw Data_Type to (Item_Code, Friendly_Name, Category, Match_Status)."""
    if not raw_data_type:
        return '', raw_data_type, '', 'UNMAPPED'
    
    raw_data_type = str(raw_data_type).strip()
    
    # Check if this is an Item_Code
    if raw_data_type in heading_by_code:
        friendly_name, category = heading_by_code[raw_data_type]
        return raw_data_type, friendly_name, category, 'EXACT'
    
    # Try exact match
    if raw_data_type in heading_lookup:
        item_code, friendly_name, category = heading_lookup[raw_data_type]
        return item_code, friendly_name, category, 'EXACT'
    
    # Fuzzy match
    import difflib
    matches = difflib.get_close_matches(raw_data_type, heading_lookup.keys(), n=1, cutoff=0.90)
    if matches:
        item_code, friendly_name, category = heading_lookup[matches[0]]
        return item_code, friendly_name, category, 'FUZZY'
    
    # UNMAPPED
    return '', raw_data_type, '', 'UNMAPPED'

def translate_financial_type(raw_type: str) -> Tuple[str, str]:
    """Translate raw Financial_Type to (Clean_Type, Match_Status)."""
    if not raw_type:
        return raw_type, 'UNMAPPED'
    
    raw_type = str(raw_type).strip()
    
    if raw_type in financial_type_lookup:
        return financial_type_lookup[raw_type], 'EXACT'
    
    import difflib
    matches = difflib.get_close_matches(raw_type, financial_type_lookup.keys(), n=1, cutoff=0.90)
    if matches:
        return financial_type_lookup[matches[0]], 'FUZZY'
    
    return raw_type, 'UNMAPPED'

# ============================================================================
# Excel Parser (simplified from financial_preprocessor_v5.py)
# ============================================================================

def extract_metadata(ws, report_year: int, report_month: int) -> List[Dict]:
    """Extract project metadata from sheet header."""
    metadata = []
    meta_fields = {
        'Project Code': 'Project Code',
        'Project Name': 'Project Name',
        'Report Date': 'Report Date',
        'Financial Status as at': 'Report Date',
        'Start Date': 'Start Date',
        'Project Start Date': 'Start Date',
        'Complete Date': 'Complete Date',
        'Project Completion Date': 'Complete Date',
        'Target Complete Date': 'Target Complete Date',
    }
    
    for row_idx in range(1, min(ws.max_row + 1, 15)):
        for label_col, value_col in [(1, 2), (2, 5)]:
            cell_label = ws.cell(row_idx, label_col).value
            cell_value = ws.cell(row_idx, value_col).value
            
            if not cell_label:
                continue
            
            label = str(cell_label).strip().rstrip(':').rstrip()
            
            if label in meta_fields:
                field_name = meta_fields[label]
                value = cell_value
                
                if value is not None:
                    if isinstance(value, str):
                        value = value.strip()
                        if '  ' in value:
                            value = value.split('  ')[0].strip()
                    else:
                        value = str(value)
                    
                    item_code, friendly_name, category, match_status = translate_data_type(field_name)
                    
                    metadata.append({
                        'project_code': None,  # Will be set later
                        'project_name': None,
                        'year': report_year,
                        'month': report_month,
                        'sheet_name': 'Financial Status',
                        'financial_type': 'General',
                        'raw_financial_type': 'General',
                        'item_code': item_code if item_code else field_name,
                        'friendly_name': friendly_name,
                        'category': category,
                        'value': value,
                        'raw_value': str(value),
                        'match_status': match_status,
                        'source_file': None,
                    })
    
    return metadata

def find_header_row(ws) -> Tuple[Optional[int], int]:
    """Find header row containing 'Item'."""
    for row_idx in range(1, min(ws.max_row + 1, 30)):
        for col in [1, 2]:
            val = ws.cell(row_idx, col).value
            if val and str(val).strip().lower() == 'item':
                return row_idx, col
    return None, 1

def get_report_date(ws) -> Tuple[int, int]:
    """Extract report date from sheet header."""
    for row_idx in range(1, 15):
        cell_a = ws.cell(row_idx, 1).value
        if cell_a and 'Report Date' in str(cell_a):
            cell_b = ws.cell(row_idx, 2).value
            if cell_b:
                date_str = str(cell_b).strip()
                if '  ' in date_str:
                    date_str = date_str.split('  ')[0].strip()
                try:
                    dt = pd.to_datetime(date_str)
                    return dt.year, dt.month
                except (ValueError, TypeError):
                    pass
    return 2025, 1

def parse_excel_file(excel_path: Path) -> List[Dict]:
    """Parse Excel file and return list of row dicts."""
    all_data = []
    
    try:
        wb = load_workbook(excel_path, data_only=True)
    except Exception as e:
        print(f"  [ERROR] Failed to load {excel_path}: {e}")
        return []
    
    # Get report date
    report_year, report_month = 2025, 1
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        report_year, report_month = get_report_date(ws)
        if report_year != 2025:
            break
    
    # Extract project code and name from filename
    filename = excel_path.stem.replace('_flat_v5', '').replace('_flat', '')
    code_match = re.match(r'^(\d+)\s*(.+)?', filename)
    project_code = code_match.group(1) if code_match else '0000'
    project_name = code_match.group(2).strip() if code_match and code_match.group(2) else filename
    
    # Parse each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        if ws.max_row < 5:
            continue
        
        header_row, item_col = find_header_row(ws)
        if header_row is None:
            continue
        
        sheet_lower = sheet_name.lower()
        is_financial_status = 'financial status' in sheet_lower
        is_monthly = any(kw in sheet_lower for kw in ['projection', 'cash flow', 'cashflow', 'committed', 'accrual', 'budget'])
        
        # Determine financial type from sheet name
        if 'cash flow' in sheet_lower or 'cashflow' in sheet_lower:
            raw_ft = 'Cash Flow'
        elif 'projection' in sheet_lower or 'projected cost' in sheet_lower:
            raw_ft = 'Projection'
        elif 'budget' in sheet_lower:
            raw_ft = 'Budget'
        elif 'committed' in sheet_lower:
            raw_ft = 'Committed Cost'
        elif 'accrual' in sheet_lower:
            raw_ft = 'Accrual'
        else:
            raw_ft = sheet_name
        
        clean_ft, ft_status = translate_financial_type(raw_ft)
        
        # Parse rows
        hierarchy_stack = []
        data_start = header_row + 4 if is_financial_status else header_row + 1
        
        # Build column mapping for Financial Status
        col_map = {}
        if is_financial_status:
            for col in range(2, ws.max_column):
                parts = []
                for row_offset in range(0, 3):
                    val = ws.cell(header_row + row_offset, col + 1).value
                    if val:
                        parts.append(str(val).strip())
                combined = ' '.join(parts)
                if combined:
                    clean_name, _ = translate_financial_type(combined)
                    if clean_name != '*not used':
                        col_map[col] = clean_name
        
        # Build month column mapping for monthly sheets
        month_map = {
            'Jan': 1, 'January': 1, 'Feb': 2, 'February': 2, 'Mar': 3, 'March': 3,
            'Apr': 4, 'April': 4, 'May': 5, 'Jun': 6, 'June': 6, 'Jul': 7, 'July': 7,
            'Aug': 8, 'August': 8, 'Sep': 9, 'September': 9, 'Oct': 10, 'October': 10,
            'Nov': 11, 'November': 11, 'Dec': 12, 'December': 12
        }
        month_cols = {}
        if not is_financial_status:
            for col in range(0, ws.max_column):
                val = ws.cell(header_row, col + 1).value
                if val:
                    if hasattr(val, 'month'):
                        month_cols[val.month] = col
                    else:
                        header = str(val).strip()
                        if header in month_map:
                            month_cols[month_map[header]] = col
        
        trade_col = item_col + 1
        
        for row_idx in range(data_start, ws.max_row + 1):
            item_code_val = ws.cell(row_idx, item_col).value
            trade_name = ws.cell(row_idx, trade_col).value
            
            if not item_code_val:
                continue
            
            item_str = str(item_code_val).strip()
            if '.' in item_str:
                parts = item_str.split('.')
                if len(parts) == 2 and parts[1] == '0':
                    item_str = parts[0]
            
            trade_str = str(trade_name).strip() if trade_name else ''
            
            if not item_str or item_str == 'None':
                continue
            
            # Build hierarchical Data_Type
            parts = item_str.replace(' ', '').split('.')
            depth = 1 if len(parts) == 2 and parts[1] == '0' else len(parts)
            while len(hierarchy_stack) >= depth:
                hierarchy_stack.pop()
            hierarchy_stack.append((item_str, trade_str))
            raw_data_type = ' - '.join([e[1] for e in hierarchy_stack if e[1]]) if any(e[1] for e in hierarchy_stack) else trade_str
            
            item_code_mapped, friendly_name, category, dt_status = translate_data_type(raw_data_type)
            
            if is_financial_status:
                # Columnar layout
                for col_idx, clean_financial_type in col_map.items():
                    if col_idx + 1 <= ws.max_column:
                        value = ws.cell(row_idx, col_idx + 1).value
                        if value is not None and value != '':
                            all_data.append({
                                'project_code': project_code,
                                'project_name': project_name,
                                'year': report_year,
                                'month': report_month,
                                'sheet_name': sheet_name,
                                'financial_type': clean_financial_type,
                                'raw_financial_type': col_map.get(col_idx, ''),
                                'item_code': item_code_mapped if item_code_mapped else item_str,
                                'friendly_name': friendly_name,
                                'category': category,
                                'value': float(value) if isinstance(value, (int, float)) else str(value),
                                'raw_value': str(value),
                                'match_status': f"{dt_status}|{ft_status}",
                                'source_file': excel_path.name,
                            })
            elif month_cols:
                # Monthly sheet with month columns
                for month_num, col_idx in month_cols.items():
                    if col_idx + 1 <= ws.max_column:
                        value = ws.cell(row_idx, col_idx + 1).value
                        if value is not None and value != '':
                            # Apply year correction
                            year = report_year if month_num <= report_month else report_year - 1
                            all_data.append({
                                'project_code': project_code,
                                'project_name': project_name,
                                'year': year,
                                'month': month_num,
                                'sheet_name': sheet_name,
                                'financial_type': clean_ft,
                                'raw_financial_type': raw_ft,
                                'item_code': item_code_mapped if item_code_mapped else item_str,
                                'friendly_name': friendly_name,
                                'category': category,
                                'value': float(value) if isinstance(value, (int, float)) else str(value),
                                'raw_value': str(value),
                                'match_status': f"{dt_status}|{ft_status}",
                                'source_file': excel_path.name,
                            })
    
    return all_data

# ============================================================================
# Supabase Upload
# ============================================================================

def upsert_project(project_code: str, project_name: str) -> str:
    """Insert or get project, return project ID."""
    # Check if exists
    result = supabase.table('projects').select('id').eq('code', project_code).execute()
    
    if result.data and len(result.data) > 0:
        return result.data[0]['id']
    
    # Insert new
    result = supabase.table('projects').insert({
        'code': project_code,
        'name': project_name
    }).execute()
    
    return result.data[0]['id']

def upload_to_supabase(rows: List[Dict], batch_size: int = 500):
    """Upload rows to Supabase in batches."""
    if not rows:
        return
    
    # Group by project
    projects: Dict[str, List[Dict]] = {}
    for row in rows:
        key = row['project_code']
        if key not in projects:
            projects[key] = []
        projects[key].append(row)
    
    total_uploaded = 0
    
    for project_code, project_rows in projects.items():
        # Get or create project
        project_name = project_rows[0]['project_name']
        project_id = upsert_project(project_code, project_name)
        print(f"  Project {project_code} ({project_name}): {len(project_rows)} rows")
        
        # Upload in batches
        for i in range(0, len(project_rows), batch_size):
            batch = project_rows[i:i + batch_size]
            
            # Prepare for insert
            insert_rows = []
            for row in batch:
                insert_rows.append({
                    'project_id': project_id,
                    'year': row['year'],
                    'month': row['month'],
                    'sheet_name': row['sheet_name'],
                    'financial_type': row['financial_type'],
                    'raw_financial_type': row['raw_financial_type'],
                    'item_code': row['item_code'],
                    'friendly_name': row['friendly_name'],
                    'category': row['category'],
                    'value': row['value'] if isinstance(row['value'], (int, float)) else None,
                    'raw_value': row['raw_value'],
                    'match_status': row['match_status'],
                    'source_file': row['source_file'],
                })
            
            try:
                result = supabase.table('financial_data').insert(insert_rows).execute()
                total_uploaded += len(insert_rows)
            except Exception as e:
                print(f"    [ERROR] Batch upload failed: {e}")
    
    print(f"[OK] Uploaded {total_uploaded} rows to Supabase")

# ============================================================================
# Main
# ============================================================================

def find_excel_files(input_dir: Path) -> List[Path]:
    """Find all Excel files in directory."""
    excel_files = []
    
    # Search recursively
    for ext in ['*.xlsx', '*.xls']:
        excel_files.extend(input_dir.rglob(ext))
    
    # Filter out _flat_v5.csv files (we want original Excel)
    excel_files = [f for f in excel_files if not str(f).endswith('.csv')]
    
    return sorted(excel_files)

def main():
    import argparse
    parser = argparse.ArgumentParser(description='Migrate Excel files to Supabase')
    parser.add_argument('--input-dir', required=True, help='Input directory with Excel files')
    parser.add_argument('--dry-run', action='store_true', help='Parse but don\'t upload')
    args = parser.parse_args()
    
    input_dir = Path(args.input_dir)
    if not input_dir.exists():
        print(f"ERROR: Directory not found: {input_dir}")
        sys.exit(1)
    
    print(f"[START] Excel to Supabase migration")
    print(f"  Input: {input_dir}")
    print(f"  Supabase: {SUPABASE_URL}")
    print(f"  Dry run: {args.dry_run}")
    
    # Load lookup tables
    load_lookup_tables()
    
    # Find Excel files
    excel_files = find_excel_files(input_dir)
    print(f"\n[INFO] Found {len(excel_files)} Excel files")
    
    if not excel_files:
        print("[WARN] No Excel files found. Make sure you're pointing to a folder with .xlsx files.")
        return
    
    # Parse and upload
    all_rows = []
    
    for i, excel_path in enumerate(excel_files, 1):
        print(f"\n[{i}/{len(excel_files)}] {excel_path.name}")
        rows = parse_excel_file(excel_path)
        print(f"  Parsed {len(rows)} rows")
        all_rows.extend(rows)
    
    print(f"\n[SUMMARY] Total: {len(all_rows)} rows from {len(excel_files)} files")
    
    if args.dry_run:
        print("[DRY RUN] Not uploading to Supabase")
    else:
        upload_to_supabase(all_rows)
    
    print("\n[DONE]")

if __name__ == '__main__':
    main()
